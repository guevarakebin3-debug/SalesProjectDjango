from io import BytesIO
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect


class SearchListMixin:
    """
    Mixin declarativo de búsqueda para cualquier ListView.

    Cada vista define search_fields como lista de dicts:
      {'param': 'q',    'fields': ['name__icontains','email__icontains']}  # OR multi-campo
      {'param': 'phone', 'field': 'phone__icontains'}                       # texto simple
      {'param': 'is_active', 'field': 'is_active', 'type': 'bool'}         # booleano
      {'param': 'price_min', 'field': 'unit_price__gte', 'type': 'number'} # numérico
      {'param': 'date_from', 'field': 'created_at__date__gte','type':'date'}# fecha

    El mixin también fija paginate_by = 10 e inyecta search_params al contexto.
    """
    search_fields = []
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset()
        for spec in self.search_fields:
            val = self.request.GET.get(spec['param'], '').strip()
            if not val:
                continue
            ftype = spec.get('type', 'text')
            if 'fields' in spec:
                q_obj = Q()
                for f in spec['fields']:
                    q_obj |= Q(**{f: val})
                qs = qs.filter(q_obj)
            elif ftype == 'bool':
                if val in ('true', 'false'):
                    qs = qs.filter(**{spec['field']: val == 'true'})
            elif ftype in ('number', 'date'):
                try:
                    qs = qs.filter(**{spec['field']: val})
                except (ValueError, TypeError):
                    pass
            else:
                qs = qs.filter(**{spec['field']: val})
        return qs.distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search_params'] = self.request.GET
        return ctx


class ExportMixin:
    """
    Mixin genérico para exportar el queryset filtrado de cualquier ListView
    a PDF o Excel.

    Uso en una vista:
        class MyListView(LoginRequiredMixin, ExportMixin, ListView):
            export_fields = [
                ('Columna', 'campo_o_callable'),
                ('Relación', 'fk__campo'),
                ('Custom', lambda obj: obj.metodo()),
            ]
            export_filename = 'mi_reporte'  # sin extensión

    Los botones en el template solo necesitan añadir ?export=pdf o ?export=excel
    a la URL actual, conservando los parámetros de búsqueda.
    """

    export_fields = []
    export_filename = None

    # ------------------------------------------------------------------ #
    # Punto de entrada: intercepta GET antes de que ListView paginee      #
    # ------------------------------------------------------------------ #
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export', '').lower()
        if fmt == 'pdf':
            return self._build_pdf()
        if fmt == 'excel':
            return self._build_excel()
        return super().get(request, *args, **kwargs)

    # ------------------------------------------------------------------ #
    # Helpers internos                                                     #
    # ------------------------------------------------------------------ #
    def _filename(self):
        if self.export_filename:
            return self.export_filename
        return getattr(getattr(self, 'model', None), '_meta', type('_', (), {'model_name': 'export'})).model_name

    def _cell(self, obj, accessor):
        """Resuelve un campo, ruta con __ o callable sobre el objeto."""
        if callable(accessor):
            return accessor(obj)
        val = obj
        for part in accessor.split('__'):
            val = getattr(val, part, '')
            if callable(val):
                val = val()
        return '' if val is None else str(val)

    def _headers(self):
        return [label for label, _ in self.export_fields]

    def _data_rows(self):
        return [
            [self._cell(obj, acc) for _, acc in self.export_fields]
            for obj in self.get_queryset()
        ]

    # ------------------------------------------------------------------ #
    # Exportación PDF                                                      #
    # ------------------------------------------------------------------ #
    def _build_pdf(self):
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        DARK  = colors.HexColor('#343a40')
        LIGHT = colors.HexColor('#f8f9fa')
        GRID  = colors.HexColor('#dee2e6')

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20,
        )

        data = [self._headers()] + self._data_rows()
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  DARK),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 8),
            ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, LIGHT]),
            ('GRID',          (0, 0), (-1, -1), 0.4, GRID),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))

        styles = getSampleStyleSheet()
        filename = self._filename()
        doc.build([Paragraph(filename.capitalize(), styles['Title']), Spacer(1, 8), table])

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response

    # ------------------------------------------------------------------ #
    # Exportación Excel                                                    #
    # ------------------------------------------------------------------ #
    def _build_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        DARK_HEX = '343A40'
        ALT_HEX  = 'F8F9FA'
        thin = Side(style='thin', color='DEE2E6')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._filename()[:31]

        ws.append(self._headers())
        for cell in ws[1]:
            cell.fill      = PatternFill('solid', fgColor=DARK_HEX)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
            cell.border    = border

        for i, row in enumerate(self._data_rows(), start=2):
            ws.append(row)
            fill = PatternFill('solid', fgColor=ALT_HEX) if i % 2 == 0 else None
            for cell in ws[i]:
                if fill:
                    cell.fill = fill
                cell.border = border

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = self._filename()
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response


class StaffRequiredMixin:
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.
    
    Uso:
        class BrandDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
            ...
    
    ¿POR QUÉ?
    Porque solo el personal autorizado (staff) debe poder
    eliminar registros. Un usuario normal puede ver y crear,
    pero no borrar información importante del sistema.
    
    ¿CÓMO FUNCIONA?
    1. El usuario intenta acceder a una vista protegida
    2. dispatch() se ejecuta ANTES que la vista
    3. Si user.is_staff es False → redirige con mensaje de error
    4. Si user.is_staff es True → ejecuta la vista normalmente
    """

    # URL a donde redirigir si no es staff
    # Se puede sobreescribir en cada vista
    staff_redirect_url = '/'
    staff_error_message = 'No tienes permiso para realizar esta acción. acceso de Staff requerido.'

    def dispatch(self, request, *args, **kwargs):
        """
        dispatch() es el primer método que se ejecuta en una CBV.
        Interceptamos aquí para verificar permisos ANTES de
        procesar la petición (GET o POST).
        """
        # Verificar si el usuario es staff
        if not request.user.is_staff:
            # Mostrar mensaje de error al usuario
            messages.error(request, self.staff_error_message)
            # Redirigir a la URL configurada
            return redirect(self.staff_redirect_url)

        # Si es staff, continuar con el flujo normal de la vista
        return super().dispatch(request, *args, **kwargs)


class SearchExportMixin(ExportMixin, SearchListMixin):
    """
    Mixin combinado listo para usar: búsqueda declarativa + paginación + PDF/Excel.

    Herencia recomendada:
        class MiListView(LoginRequiredMixin, SearchExportMixin, ListView):
            search_fields = [...]
            export_fields = [...]
            export_filename = 'mi_reporte'
    """
    pass
